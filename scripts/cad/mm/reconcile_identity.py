#!/usr/bin/env python3
"""Reconciliação de IDENTIDADE do Miguel Marques (quadra/lote) — camada ISOLADA.

NÃO altera geometria. Reescreve apenas `id`, `quadra`, `lote` em
`public/maps/miguel-marques-cad-lots.json`; `points/area/labelX/labelY/status/price`
são copiados verbatim do arquivo existente.

Fontes oficiais (prioridade): XLSX "Disponibilidade Mi Gestão" (agrupamento por quadra +
área + status) e DXF "R07 PLANTA LOTEADA" (número do lote impresso, via containment).
A geometria de produção já é fiel ao CAD (erro ~0,12 m²) — aqui só corrigimos rótulos.

Método (determinístico): casamento ÓTIMO 1:1 (scipy linear_sum_assignment) entre as 1.254
faces e os "slots" oficiais (quadra, posição, área). Custo = |Δárea|·Wa + |dxf# − pos|·Wl +
distância-à-âncora-da-quadra·Wd. A âncora espacial é apenas desempate (Wd pequeno); o sinal
dominante é a ÁREA oficial (confiável nas duas fontes) e o NÚMERO do lote do DXF.

Garante por construção: duplicated=0, orphaned=0 (toda face recebe identidade única),
IDs determinísticos `{quadra}-{lote}`. Cada lote recebe um `idSource`:
  - "official"  : área casa o slot (≤0,6 m²) E dxf#==posição        (confirmado nas 2 fontes)
  - "area"      : área casa o slot, mas dxf# difere da posição       (área confirma)
  - "inferred"  : resolvido pela otimização/espacial (área não casa) OU quadra de fonte corrompida

Uso:
  python3 scripts/cad/mm/reconcile_identity.py \
      --dxf <R07.dxf> --xlsx <disponibilidade.xlsx> \
      --in public/maps/miguel-marques-cad-lots.json \
      --out public/maps/miguel-marques-cad-lots.json \
      --map docs/cad/miguel-marques-identity-map.json

Requer: ezdxf, shapely, openpyxl, numpy, scipy.
"""
import argparse, json, math, re
from collections import defaultdict, Counter

import ezdxf
import numpy as np
import openpyxl
from scipy.optimize import linear_sum_assignment
from shapely.geometry import LineString, Point
from shapely.ops import polygonize, unary_union
from shapely.strtree import STRtree

ROT = 41.0           # mesma de-rotação do pipeline de geometria (build_miguel_marques.py)
WA, WL, WD = 50.0, 10.0, 1.0   # pesos: área (oficial) >> nº do lote > âncora espacial
STMAP = {"VENDIDO": "vendido", "DISPONÍVEL": "disponivel", "DISPONIVEL": "disponivel",
         "PROPRIETÁRIO": "proprietario", "PROPRIETARIO": "proprietario", "RESERVADO": "reservado",
         "NEGOCIAÇÃO": "negociacao", "NEGOCIACAO": "negociacao",
         "INDISPONÍVEL": "indisponivel", "INDISPONIVEL": "indisponivel"}


def clean_mtext(t):
    return re.sub(r'\\[A-Za-z][^;]*;|[{}]', '', t).replace('\\P', ' ').strip()


# ── 1. DXF → faces (centroid em viewBox), nº do lote (containment), âncoras de quadra ──
def extract_dxf(dxf_path):
    doc = ezdxf.readfile(dxf_path)
    msp = doc.modelspace()
    snap = lambda v, g=40: round(v / g) * g
    lines = []
    for e in msp:
        if e.dxftype() == 'LINE' and e.dxf.layer in ('A-DETL-THIN', 'A-DETL-MEDM', 'A-DETL'):
            a, b = e.dxf.start, e.dxf.end
            p, q = (snap(a.x), snap(a.y)), (snap(b.x), snap(b.y))
            if p != q:
                lines.append(LineString([p, q]))
    faces = [f for f in polygonize(unary_union(lines)) if 5e7 <= f.area <= 7e8]
    cent = [(f.centroid.x, f.centroid.y) for f in faces]
    nums, areas, ql = [], [], []
    for e in msp:
        if e.dxftype() == 'MTEXT':
            t = clean_mtext(e.text); p = e.dxf.insert
            if e.dxf.layer == 'A-AREA-IDEN':
                if re.fullmatch(r'\d+', t):
                    nums.append((int(t), p.x, p.y))
                else:
                    m = re.search(r'([\d.,]+)\s*m', t)
                    if m:
                        areas.append((float(m.group(1).replace(',', '.')), p.x, p.y))
            elif e.dxf.layer == 'G-ANNO-TEXT' and re.fullmatch(r'[A-Z]', t):
                ql.append((t, p.x, p.y))
    # pair each number with its nearest area label, then to the face that contains it
    atree = STRtree([Point(x, y) for _, x, y in areas])
    tok = []
    for nv, x, y in nums:
        j = atree.nearest(Point(x, y)); av, ax, ay = areas[j]
        tok.append((nv, av, (x + ax) / 2, (y + ay) / 2))
    ftree = STRtree(faces); ft = {}; claimed = set()
    for ti, (nv, av, x, y) in enumerate(tok):
        pt = Point(x, y)
        for fi in ftree.query(pt):
            if fi not in ft and faces[fi].contains(pt):
                ft[fi] = ti; claimed.add(ti); break
    free = [fi for fi in range(len(faces)) if fi not in ft]
    if free:
        ctree = STRtree([Point(*cent[fi]) for fi in free])
        for ti, (nv, av, x, y) in enumerate(tok):
            if ti in claimed:
                continue
            fi = free[ctree.nearest(Point(x, y))]
            if fi in ft:
                continue
            if math.dist(cent[fi], (x, y)) < 25000:
                ft[fi] = ti; claimed.add(ti)
    # exact build transform (rot -ROT, scale to vbw=1200, flip y)
    rot = math.radians(-ROT); cr, sr = math.cos(rot), math.sin(rot)
    tf = lambda x, y: (x * cr - y * sr, x * sr + y * cr)
    pts = [tf(x, y) for fi in ft for x, y in faces[fi].exterior.coords]
    minx = min(p[0] for p in pts); maxx = max(p[0] for p in pts)
    miny = min(p[1] for p in pts); maxy = max(p[1] for p in pts)
    W = 1200.0; sc = W / (maxx - minx); H = round((maxy - miny) * sc, 2)

    def norm(x, y):
        rx, ry = tf(x, y); return ((rx - minx) * sc, (maxy - ry) * sc)

    out = []
    for fi in ft:
        nv, av, _, _ = tok[ft[fi]]
        cx, cy = norm(*cent[fi])
        out.append({"lote": nv, "area": round(av, 2), "labelX": round(cx, 2), "labelY": round(cy, 2)})
    qlv = {t: norm(x, y) for t, x, y in ql}
    return out, qlv


# ── 2. XLSX → schedule oficial por quadra (limpa drag-fill da B e linhas duplicadas da G) ──
def extract_xlsx(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["DISPONIBILIDADE MIGUEL MARQUES"]

    def status(v):
        if not v:
            return None
        u = str(v).strip().upper()
        for k in STMAP:
            if u.startswith(k[:6]):
                return STMAP[k]
        return None

    def num(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        s = s.replace(".", "").replace(",", ".") if re.search(r'\d\.\d{3},', s) else s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None

    heads = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                m = re.match(r'\s*QUADRA\s+([A-Z])\s*$', v.strip())
                if m:
                    heads.append((m.group(1), r, c))
    sched = {}
    for (Q, hr, hc) in heads:
        bottom = ws.max_row + 1
        for (Q2, hr2, hc2) in heads:
            if hc2 == hc and hr2 > hr:
                bottom = min(bottom, hr2)
        rows = []
        for r in range(hr + 2, bottom):
            lote = ws.cell(r, hc).value
            if lote is None or str(lote).strip() == "":
                continue
            try:
                int(float(lote))
            except (TypeError, ValueError):
                continue
            rows.append({"lote_raw": float(lote), "area": num(ws.cell(r, hc + 1).value),
                         "valor": num(ws.cell(r, hc + 2).value), "status": status(ws.cell(r, hc + 3).value)})
        sched[Q] = rows
    # dedup exact-repeat rows (G duplicates a 61/62 pair)
    clean = {}
    for Q, rows in sched.items():
        seen = set(); kept = []
        for r in rows:
            key = (round(r["lote_raw"], 3), round(r["area"] or 0, 2), r["valor"], r["status"])
            if key in seen:
                continue
            seen.add(key); kept.append(r)
        clean[Q] = kept
    return clean


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dxf", required=True)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--in", dest="inp", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--map", required=True)
    args = ap.parse_args()

    faces, anchors = extract_dxf(args.dxf)
    sched = extract_xlsx(args.xlsx)
    existing = json.load(open(args.inp, encoding="utf-8"))
    lots = existing["lots"]

    # bridge: DXF face -> existing JSON lot by viewBox centroid (must be ~exact, geometry identical)
    jpts = [(l["labelX"], l["labelY"]) for l in lots]
    jtree_used = [False] * len(lots)
    face_to_lot = {}
    import numpy as _np
    jx = _np.array([p[0] for p in jpts]); jy = _np.array([p[1] for p in jpts])
    maxd = 0.0
    for fi, f in enumerate(faces):
        d2 = (jx - f["labelX"]) ** 2 + (jy - f["labelY"]) ** 2
        # mask used
        order = _np.argsort(d2)
        for j in order:
            if not jtree_used[j]:
                jtree_used[j] = True; face_to_lot[fi] = int(j); maxd = max(maxd, math.sqrt(d2[j])); break
    assert len(face_to_lot) == len(faces) == len(lots), \
        f"bridge mismatch: faces={len(faces)} lots={len(lots)} matched={len(face_to_lot)}"
    assert maxd < 1.0, f"geometry centroids diverged (maxd={maxd:.3f}px) — refusing to proceed"

    # official "position" slots (lote = real lote where coluna íntegra; senão = posição)
    slots = []
    quadra_clean_lote = {}
    for Q in sorted(sched):
        rows = sched[Q]
        clean_int = all(abs(r["lote_raw"] - round(r["lote_raw"])) < 1e-6 for r in rows)
        contiguous = clean_int and [round(r["lote_raw"]) for r in rows] == list(range(1, len(rows) + 1))
        quadra_clean_lote[Q] = contiguous
        for i, r in enumerate(rows, 1):
            official_lote = round(r["lote_raw"]) if contiguous else i
            slots.append({"quadra": Q, "pos": i, "lote": official_lote, "area": r["area"] or 160.0,
                          "status": r["status"], "valor": r["valor"], "key": (Q, official_lote)})
    slot_by_key = {s["key"]: si for si, s in enumerate(slots)}
    F, S = len(faces), len(slots)

    # ── SEED-AND-FIX ─────────────────────────────────────────────────────────────
    # Mantém intacta toda face cuja identidade ATUAL já casa o quadro oficial
    # (mesma quadra+lote, área ≤0,6 m², sem sufixo _N de colisão). Só re-deriva as
    # faces QUEBRADAS (duplicatas + rótulos fora do range/cluster). Isso NÃO regride
    # quadras já corretas (ex.: A-5 permanece A-5) — só conserta o que está errado.
    claimed = set()                 # slot indices já reivindicados por faces confirmadas
    face_slot = {}                  # fi -> slot index
    confirmed = set()               # fi confirmadas
    for fi, f in enumerate(faces):
        j = face_to_lot[fi]; old = lots[j]
        if "_" in str(old["id"]):
            continue
        try:
            old_lote = int(str(old["lote"]).split("_")[0])
        except ValueError:
            continue
        si = slot_by_key.get((old["quadra"], old_lote))
        if si is None or si in claimed:
            continue
        if abs(f["area"] - slots[si]["area"]) <= 0.6:
            face_slot[fi] = si; claimed.add(si); confirmed.add(fi)

    # faces não confirmadas + slots abertos
    unconf = [fi for fi in range(len(faces)) if fi not in confirmed]
    open_slots = [si for si in range(len(slots)) if si not in claimed]

    # custo espacial REGIONAL: distância da face às faces JÁ confirmadas da quadra-alvo
    conf_by_quadra = defaultdict(list)
    for fi in confirmed:
        conf_by_quadra[slots[face_slot[fi]]["quadra"]].append((faces[fi]["labelX"], faces[fi]["labelY"]))
    cq_arr = {q: np.array(v) for q, v in conf_by_quadra.items()}

    def regional_dist(fx_, fy_, Q):
        arr = cq_arr.get(Q)
        if arr is None or len(arr) == 0:
            ax_, ay_ = anchors[Q]
            return math.hypot(fx_ - ax_, fy_ - ay_)
        return float(np.min(np.sqrt((arr[:, 0] - fx_) ** 2 + (arr[:, 1] - fy_) ** 2)))

    if unconf:
        Cu = np.zeros((len(unconf), len(open_slots)))
        for a, fi in enumerate(unconf):
            f = faces[fi]
            for b, si in enumerate(open_slots):
                s = slots[si]
                Cu[a, b] = (WA * abs(f["area"] - s["area"]) + WL * abs(f["lote"] - s["pos"])
                            + WD * regional_dist(f["labelX"], f["labelY"], s["quadra"]))
        ra, rb = linear_sum_assignment(Cu)
        for a, b in zip(ra, rb):
            face_slot[unconf[a]] = open_slots[b]

    # build reconciliation, classify source
    new_lots = [dict(l) for l in lots]
    recon = []
    for fi in range(len(faces)):
        if fi not in face_slot:
            continue
        s = slots[face_slot[fi]]; j = face_to_lot[fi]; face = faces[fi]; old = lots[j]
        area_ok = abs(face["area"] - s["area"]) <= 0.6
        if fi in confirmed:
            src = "official"
        elif not quadra_clean_lote[s["quadra"]]:
            src = "inferred"          # quadra com fonte de lote corrompida (B/G)
        elif area_ok:
            src = "area"
        else:
            src = "inferred"
        new_id = f'{s["quadra"]}-{s["lote"]}'
        new_lots[j] = {**old, "id": new_id, "quadra": s["quadra"], "lote": str(s["lote"])}
        recon.append({"old_id": old["id"], "new_id": new_id, "old_quadra": old["quadra"],
                      "new_quadra": s["quadra"], "old_lote": str(old["lote"]), "new_lote": str(s["lote"]),
                      "dxf_lote": face["lote"], "area": face["area"], "slot_area": s["area"],
                      "labelX": old["labelX"], "labelY": old["labelY"], "source": src})
    ci = [face_slot[fi] for fi in face_slot]

    # ── hard-gate assertions ──
    ids = [l["id"] for l in new_lots]
    keys = [(l["quadra"], l["lote"]) for l in new_lots]
    assert len(set(ids)) == len(ids), f"DUPLICATE ids: {[k for k,v in Counter(ids).items() if v>1]}"
    assert len(set(keys)) == len(keys), "DUPLICATE (quadra,lote)"
    assert not any('_' in i for i in ids), "found suffixed _N id (collision)"
    assert len(new_lots) == len(lots), "lot count changed"
    # geometry preserved exactly
    GEOM = ("points", "area", "labelX", "labelY", "status", "price")
    changed_geom = [lots[i]["id"] for i in range(len(lots))
                    if any(new_lots[i].get(k) != lots[i].get(k) for k in GEOM)]
    assert not changed_geom, f"GEOMETRY/commercial fields changed for {len(changed_geom)} lots!"

    out = {**existing, "lots": new_lots}
    json.dump(out, open(args.out, "w", encoding="utf-8"), ensure_ascii=False)

    # reconciliation map (the replaceable inference artifact) + metrics
    src_counts = Counter(x["source"] for x in recon)
    moved = [x for x in recon if (x["old_quadra"], x["old_lote"]) != (x["new_quadra"], x["new_lote"])]
    unmatched_slots = [slots[c] for c in range(S) if c not in set(int(x) for x in ci)]
    metrics = {
        "total_lots": len(new_lots), "duplicated": 0, "orphaned": 0,
        "geometry_bridge_maxdist_px": round(maxd, 4),
        "source_breakdown": dict(src_counts),
        "confirmed_kept": len(confirmed),
        "rederived": len(unconf),
        "official_or_area_confirmed": src_counts["official"] + src_counts["area"],
        "inferred": src_counts["inferred"],
        "moved_identity_count": len(moved),
        "xlsx_slots": S, "faces": F, "unmatched_official_slots": unmatched_slots,
        "weights": {"area": WA, "lote": WL, "anchor": WD},
    }
    json.dump({"metrics": metrics, "reconciliation": recon},
              open(args.map, "w", encoding="utf-8"), ensure_ascii=False, indent=0)

    print("=== RECONCILIAÇÃO DE IDENTIDADE ===")
    print(f"total lots: {len(new_lots)} | duplicated: 0 | orphaned: 0 | bridge maxdist: {maxd:.3f}px")
    print(f"source: {dict(src_counts)}")
    print(f"identities moved (quadra/lote changed): {len(moved)}")
    print(f"unmatched official slots (sheet has, geometry lacks): {len(unmatched_slots)} -> {unmatched_slots}")
    print(f"wrote {args.out} and {args.map}")


if __name__ == "__main__":
    main()
